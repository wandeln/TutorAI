"""
Excel-Export-Service: Generiert .xlsx-Dateien aus Übersichtstabellen.

Exportiert alle Students × Tasks mit erreichten Punkten + Summen.
Summe und Anteil werden als Excel-Formeln berechnet, sodass man
nachträglich noch Korrekturen in den Zellen vornehmen kann.

Unterstützt Filter nach Blatt/Typ und Styling.
"""

import io
from typing import Optional
from sqlmodel import Session
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from config import EXCEL_SHEET_NAME
from database.models import Task, User


class ExportService:
    """
    Generiert Excel-Exporte für Tutor-Übersichten.

    Usage:
        export = ExportService()
        workbook = export.generate_overview(course_id, students, tasks, scores, session)
        return StreamingResponse(io.BytesIO(workbook.read()), media_type=..., filename=...)
    """

    # Styling-Constants
    HEADER_FONT = Font(name="Calibri", bold=True, size=12, color="FFFFFF")
    HEADER_FILL = PatternFill(start_color="2E74B5", end_color="2E74B5", fill_type="solid")
    SUM_FONT = Font(name="Calibri", bold=True, size=11)
    PASS_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Grün
    FAIL_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # Rot
    THIN_BORDER = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    def generate_overview(
        self,
        course_name: str,
        students: list[User],
        tasks: list[Task],
        scores: dict,  # {student_id: {task_id: points}}
        session: Session,
        filter_text: Optional[str] = None,
    ) -> Workbook:
        """
        Generiert eine Excel-Übersicht.

        Structure:
          Row 1: Course header (merged)
          Row 2: Headers (Name, Matr.-Nr., Task1, Task2, ..., Summe, Anteil)
          Row 3+: Data (name, matrnr, points, ..., SUM-formula, %-formula)

        Summe und Anteil werden als Excel-Formeln berechnet!
        """
        wb = Workbook()
        ws = wb.active
        assert ws is not None
        ws.title = EXCEL_SHEET_NAME

        from openpyxl.utils import get_column_letter

        # ─── Filter Tasks wenn nötig ────────────────────────────
        if filter_text:
            filter_lower = filter_text.lower()
            tasks = [t for t in tasks if filter_lower in t.title.lower()]

        num_tasks = len(tasks)

        # Spalten-Indizes (1-basiert)
        COL_NAME = 1
        COL_MATRNR = 2
        COL_FIRST_TASK = 3
        COL_SUM = COL_FIRST_TASK + num_tasks
        COL_PCT = COL_SUM + 1

        # ─── Row 1: Kurs-Header ─────────────────────────────────
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=COL_PCT)
        header_cell = ws.cell(row=1, column=1, value=f"Punktestand: {course_name}")
        header_cell.font = Font(name="Calibri", bold=True, size=14, color="1F4E79")
        header_cell.alignment = Alignment(horizontal="center")

        # ─── Row 2: Spalten-Header ──────────────────────────────
        header_fill = self.HEADER_FILL
        header_font = self.HEADER_FONT

        headers = ["Name", "Matr.-Nr."]
        for task in tasks:
            headers.append(task.title)
        headers.extend(["Summe", "Anteil (%)"])

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
            cell.border = self.THIN_BORDER

        # ─── Rows 3+: Daten ─────────────────────────────────────
        max_possible = sum(t.max_points for t in tasks)

        for row_idx, student in enumerate(students, 3):
            student_id = student.id
            student_scores = scores.get(student_id, {})

            # Name
            ws.cell(row=row_idx, column=COL_NAME, value=student.name).font = Font(
                name="Calibri"
            )
            ws.cell(row=row_idx, column=COL_NAME).border = self.THIN_BORDER

            # Matr.-Nr. (username als Ersatz)
            ws.cell(row=row_idx, column=COL_MATRNR, value=student.username).font = Font(
                name="Calibri"
            )
            ws.cell(row=row_idx, column=COL_MATRNR).border = self.THIN_BORDER

            # Einzelpunkte pro Task
            for col, task in enumerate(tasks, COL_FIRST_TASK):
                points = student_scores.get(task.id, 0)
                cell = ws.cell(row=row_idx, column=col, value=points)
                cell.alignment = Alignment(horizontal="center")
                cell.border = self.THIN_BORDER

                # Bedingte Formatierung: Grün/Rot basierend auf %
                if task.max_points > 0:
                    pct = points / task.max_points
                    cell.fill = self.PASS_FILL if pct >= 0.5 else self.FAIL_FILL

            # Summe als Excel-Formel: =SUM(C{row}:D{row})
            sum_cell = ws.cell(row=row_idx, column=COL_SUM)
            task_range_start = get_column_letter(COL_FIRST_TASK)
            task_range_end = get_column_letter(COL_FIRST_TASK + num_tasks - 1)
            sum_cell.value = f"=SUM({task_range_start}{row_idx}:{task_range_end}{row_idx})"
            sum_cell.font = self.SUM_FONT
            sum_cell.border = self.THIN_BORDER
            sum_cell.alignment = Alignment(horizontal="center")

            # Anteil % als Excel-Formel: =IF(max_possible>0, Summe/max_possible*100, 0)
            pct_cell = ws.cell(row=row_idx, column=COL_PCT)
            sum_col_letter = get_column_letter(COL_SUM)
            if max_possible > 0:
                pct_cell.value = (
                    f'=IF({sum_col_letter}{row_idx}="",'
                    f"""'',TEXT({sum_col_letter}{row_idx}/{max_possible},"0.0%"))"""
                )
            else:
                pct_cell.value = '="0.0%"'
            pct_cell.font = self.SUM_FONT
            pct_cell.alignment = Alignment(horizontal="center")
            pct_cell.border = self.THIN_BORDER

        # ─── Spaltenbreiten anpassen ────────────────────────────
        ws.column_dimensions["A"].width = 20  # Name
        ws.column_dimensions["B"].width = 15  # Matr.-Nr.
        for col in range(COL_FIRST_TASK, COL_FIRST_TASK + num_tasks):
            col_letter = get_column_letter(col)
            ws.column_dimensions[col_letter].width = 25  # Tasks
        ws.column_dimensions[get_column_letter(COL_SUM)].width = 12  # Summe
        ws.column_dimensions[get_column_letter(COL_PCT)].width = 12  # %

        return wb

    def generate_overview_bytes(
        self,
        course_name: str,
        students: list[User],
        tasks: list[Task],
        scores: dict,
        session: Session,
        filter_text: Optional[str] = None,
    ) -> bytes:
        """Wie generate_overview, aber gibt Bytes zurück (für Response)."""
        wb = self.generate_overview(
            course_name, students, tasks, scores, session, filter_text
        )
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()